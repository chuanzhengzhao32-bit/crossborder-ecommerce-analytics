"""Stream-clean the UCI workbook into an analysis-ready CSV and profile."""

from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "raw" / "uci_online_retail_ii" / "online_retail_II.xlsx"
OUTPUT_DIR = ROOT / "data" / "processed"
OUTPUT = OUTPUT_DIR / "fact_order_lines.csv"
PROFILE = OUTPUT_DIR / "source_profile.json"
PROFILE_DOC = ROOT / "docs" / "source_profile.json"

OUTPUT_FIELDS = [
    "transaction_line_id",
    "invoice_no",
    "stock_code",
    "description",
    "quantity",
    "invoice_datetime",
    "invoice_date",
    "unit_price_gbp",
    "line_revenue_gbp",
    "customer_id",
    "country",
    "is_cancellation",
    "is_return_line",
    "is_valid_sale",
    "source_sheet",
]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def clean_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(
            f"Missing {SOURCE}. Run scripts/download_source_data.py first."
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)

    row_count = 0
    usable_row_count = 0
    cancellation_count = 0
    return_line_count = 0
    valid_sale_count = 0
    missing_customer_count = 0
    countries: Counter[str] = Counter()
    stock_codes: set[str] = set()
    min_date: datetime | None = None
    max_date: datetime | None = None

    with OUTPUT.open("w", newline="", encoding="utf-8-sig") as target:
        writer = csv.DictWriter(target, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()

        line_id = 0
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(min_row=2, values_only=True)
            for invoice, stock_code, description, quantity, invoice_date, price, customer_id, country in rows:
                row_count += 1
                if invoice is None or stock_code is None or invoice_date is None:
                    continue

                invoice_no = clean_id(invoice)
                stock_code_text = clean_id(stock_code)
                quantity_number = int(quantity or 0)
                price_number = float(price or 0)
                customer_text = clean_id(customer_id)
                country_text = clean_text(country) or "Unknown"

                is_cancellation = invoice_no.upper().startswith("C")
                is_return_line = quantity_number < 0 or is_cancellation
                is_valid_sale = (
                    quantity_number > 0
                    and price_number > 0
                    and not is_cancellation
                )

                line_id += 1
                usable_row_count += 1
                cancellation_count += int(is_cancellation)
                return_line_count += int(is_return_line)
                valid_sale_count += int(is_valid_sale)
                missing_customer_count += int(not customer_text)
                countries[country_text] += 1
                stock_codes.add(stock_code_text)
                min_date = invoice_date if min_date is None else min(min_date, invoice_date)
                max_date = invoice_date if max_date is None else max(max_date, invoice_date)

                writer.writerow(
                    {
                        "transaction_line_id": line_id,
                        "invoice_no": invoice_no,
                        "stock_code": stock_code_text,
                        "description": clean_text(description),
                        "quantity": quantity_number,
                        "invoice_datetime": invoice_date.isoformat(sep=" "),
                        "invoice_date": invoice_date.date().isoformat(),
                        "unit_price_gbp": round(price_number, 4),
                        "line_revenue_gbp": round(quantity_number * price_number, 4),
                        "customer_id": customer_text,
                        "country": country_text,
                        "is_cancellation": int(is_cancellation),
                        "is_return_line": int(is_return_line),
                        "is_valid_sale": int(is_valid_sale),
                        "source_sheet": sheet.title,
                    }
                )

    profile = {
        "dataset": "UCI Online Retail II",
        "source_workbook": str(SOURCE.relative_to(ROOT)),
        "source_sheets": workbook.sheetnames,
        "raw_rows_seen": row_count,
        "usable_rows_written": usable_row_count,
        "valid_sale_lines": valid_sale_count,
        "cancellation_lines": cancellation_count,
        "return_lines": return_line_count,
        "missing_customer_lines": missing_customer_count,
        "date_min": min_date.isoformat() if min_date else None,
        "date_max": max_date.isoformat() if max_date else None,
        "distinct_countries": len(countries),
        "distinct_stock_codes": len(stock_codes),
        "top_countries_by_lines": countries.most_common(15),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    profile_text = json.dumps(profile, ensure_ascii=False, indent=2)
    PROFILE.write_text(profile_text, encoding="utf-8")
    PROFILE_DOC.write_text(profile_text, encoding="utf-8")
    print(json.dumps(profile, ensure_ascii=False, indent=2))
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
